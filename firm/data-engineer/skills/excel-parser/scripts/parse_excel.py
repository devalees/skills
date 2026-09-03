import os
import sys
import json
import openpyxl

def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_data = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            continue

        non_empty_rows = [list(r) for r in raw_rows if any(cell is not None and str(cell).strip() != '' for cell in r)]
        if not non_empty_rows:
            continue

        header_idx = 0
        for i, row in enumerate(non_empty_rows):
            valid_cells = [c for c in row if c is not None and str(c).strip() != '']
            if len(valid_cells) >= 2:
                header_idx = i
                break

        raw_headers = non_empty_rows[header_idx]
        headers = []
        for col_idx, h in enumerate(raw_headers):
            val = str(h).strip() if h is not None else f"col_{col_idx+1}"
            headers.append(val)

        data_rows = []
        for r in non_empty_rows[header_idx + 1:]:
            row_dict = {}
            for col_idx, col_name in enumerate(headers):
                cell_val = r[col_idx] if col_idx < len(r) else None
                row_dict[col_name] = cell_val
            data_rows.append(row_dict)

        total_rows += len(data_rows)
        sheets_data.append({
            "sheet_name": sheet_name,
            "detected_header_row": header_idx + 1,
            "headers": headers,
            "row_count": len(data_rows),
            "rows": data_rows
        })

    envelope = {
        "document_metadata": {
            "file_name": os.path.basename(file_path),
            "file_type": "excel",
            "structure_type": "tabular",
            "sheet_count": len(sheets_data)
        },
        "payload": {
            "tables": sheets_data
        },
        "data_health": {
            "total_rows_extracted": total_rows,
            "sheets_processed": len(sheets_data),
            "status": "success"
        }
    }
    return envelope

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 parse_excel.py <path_to_excel_file>")
        sys.exit(1)
    result = parse_excel(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))
